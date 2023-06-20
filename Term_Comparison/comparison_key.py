from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

"""
This class represents the bank of words that will be tested against.
An instantiation of this class will contain a dictionary with every term from the textbook designated
as the key to compare terms from other textbooks against. These terms will be orgnanized by chapter
and this class contains methods to compare terms from other textbooks against the key.
The textbook used as the key is determined by the textbook in the comparison sheet that is in Column B.
"""
class comparisonKey:
  """
  file_name: the name of the sheet designated as the key
  """
  def __init__(self, file_name: str) -> None:
    # Class variable to hold all terms from sheet designated as the comparison key
    wb: Workbook = load_workbook(file_name)
    self.key_textbook_name: str  = file_name[file_name.rfind("/")+1:file_name.find("xlsx")-1]
    self.key = self.__intialize_key(wb)

  """
  file_name: the name of the sheet designated as the key
  Reads the sheets from wb to create the key to compare against
  """
  def __intialize_key(self, wb: Workbook) -> dict[(str, str), list[str]]:
    key: dict[(str, str), list[str]] = {}
    for sheet_name in wb.sheetnames:
      sheet: Worksheet = wb[sheet_name]
      chapter_name: str = sheet['A1'].value
      terms: list[str] = []
      for value in sheet.iter_cols(min_row=3, values_only=True):
        for term in value:
          if term != None:
            terms.append(term)
      key.update({(sheet_name, chapter_name):terms})
    return key
  
  """
  folder_path: path to OneDrive folder containing spreadsheets
  key_chapter: tuple of chapter acting as key for comparison
  textbooks: List of all textbooks being compared against the ky
  general_chapter: General chapter that current chapter of slef.key corresponds to
  textbook_index: Corresponding index of current textbook (see more in word_count_tracker NOTE in compare_main.py)
  chapter: Name and sheet name of chapter being compared
  word_count_tracker: data structure tracking common terms (see more in word_count_tracker NOTE in compare_main.py)
  Compares values from the appropriate chapter in self.key to the values in the chapter being passed in and updates
    and returns word_count_tracker accordingly 
  """
  def compare_textbook_chapter(self, folder_path: str, key_chapter: tuple[str, str], textbooks: list[str], general_chapter: str, textbook_index: int, chapter: tuple[str, str], word_count_tracker: dict[tuple[str, str], list[int]]) -> dict[tuple[str, str], list[int]]:
    sheet_name, chapter_name = chapter
    wb: Workbook = load_workbook(f"{folder_path}{textbooks[textbook_index]}.xlsx")
    ws: Worksheet = wb[sheet_name]
    for value in ws.iter_cols(min_row=3, values_only=True):
      for term in value:
        if term in self.key.get(key_chapter):
          if (general_chapter, term) in word_count_tracker.keys():
            list_count = word_count_tracker.get((general_chapter, term))
          else:
            list_count = [0] * len(textbooks)
          list_count[textbook_index] = 1
          word_count_tracker.update({(general_chapter, term):list_count})
    return word_count_tracker