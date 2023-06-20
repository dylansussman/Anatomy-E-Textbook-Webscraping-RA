from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

"""
NOTE
Only things that may need to change are the values of the min/max_row/col in iter_row/col depending on which textbook is used
as the key and whether more textbooks are added to compare
"""
"""
This class contains all the mappings needed to compare terms to a key and output the result to a sheet
"""
class comparisonMappings:
  """
  file_name: the name of the sheet containing the info to create necessary mappings
  self.general_chapters: mapping from title of each chapter from the textbook used as the comparison key
    to the corresponding general "chapter" name
    Done this way (textbook chapter title (key) -> general "chapter" title (value)) because each textbook chapter corresponds
    to one and only one general chapter, but general chapters can correspond to > 1 chapter
    The first tuple entry corresponds to the sheet name for that chapter data, the second entry corresponds to the textbook title of that chapter
  self.textbook_titles: list containing the titles of all the textbooks except the one being used as the key, in the order they appear in the spreadsheet
    The title of each textbook in this sheet should match excatly to the names of the corresponding Excel Workbooks that contain all of that textbook's
    bolded words. 
  self.chapters_key: mapping from title of each key chapter to the corresponding chapters in the other textbooks
  """
  def __init__(self, file_name: str) -> None:
    wb: Workbook = load_workbook(file_name)
    ws: Worksheet = wb["Comparison Key"] # NOTE If the name of the sheet in Textbook Chapter Comparisons changes, change this to correspond to that
    self.general_chapters = self.__initialize_general_chapter_map(ws)
    self.textbook_titles, self.chapters_key = self.__initialize_key_chapter_map(ws)
    self.key_textbook: str = self.__initialize_key_textbook(ws)

  def __initialize_key_textbook(self, ws: Worksheet) -> str:
    return ws.cell(row=1, column=2).value
  
  """
  wb: the sheet containing the info to create necessary mappings
  Reads in appropriate columns of the sheet to create a mapping that relates
  each chapter from the textbook being used as the comparison key to the general chapter names
  and returns that mapping
  """
  def __initialize_general_chapter_map(self, ws: Worksheet) -> dict[(str, str), str]:
    key: dict[(str, str), str] = {}
    current_general_chapter: str = ""
    for value in ws.iter_rows(min_row=2, max_row=34, min_col=1, max_col=2, values_only=True):
      general_title, textbook_title = value
      if textbook_title != None:
        chapter_tuple = self.__get_chapter_tuple(textbook_title)
        if general_title != None:
          current_general_chapter = general_title
        key.update({chapter_tuple:current_general_chapter})
    return key

  """
  NOTE Dealing with merged cells and empty cells
  For key, keep track of current_key_chapter like in __initialize_general_chapter_map
  For values, if a chapter key is not None, but a corresponding chapter is, then get the list corresponding
  to that current key and use textbook_titles to find what the chapter should be (e.g., Chapter 2 from Gartner and Hiatt's)
  If a cell has the value 'Empty' then there is no corresponding chapter so make the entry in the list None (this should be the only time
  where list entries are None)
  If the value of the first column (chapter key) is None, then all entries from that row that are not None 
  need to be added to that key's list in the following way:
    The tuple of the corresponding entry in the list should be changed to also include the new tuple, thus each key's list
    should always have a length of 8 (or whatever the number of textbook is) and the tuples for each entry in the list will contain
    all the chapters from that textbook that match the key's chapter (i.e., tuples will always have a length that is a multiple of 2)
  If the value of the first column isn't None, do what is described above
  """
  """
  wb: the sheet containing the info to create necessary mappings
  Reads in appropriate columns of the sheet to create a mapping that relates each
  key chapter (chapter titles corresponding to the textbook being used as the comparison key)
  to the matching chapters (based on topic, specified in spreadsheet) from all other textbooks being compared
  and returns that mapping in addition to returning a list of all the textbook titles except the one being used
  as the key
  """
  def __initialize_key_chapter_map(self, ws: Worksheet) -> tuple[list[str], dict[(str, str), list[(str, str)]]]:
    key: dict[(str, str), list[(str, str)]] = {}
    textbook_titles: list = []
    for value in ws.iter_cols(min_row=1, max_row=1, min_col=3, max_col=10, values_only=True):
      title, = value
      textbook_titles.append(title)
    current_key_chapter: str = ""
    for value in ws.iter_rows(min_row=2, max_row=34, min_col=2, max_col=10, values_only=True):
      chapters = list(value)
      key_chapter = chapters.pop(0)
      chapter_list: list[(str, str)] = []
      if key_chapter == None:
        chapter_list = key.get(current_key_chapter)
        for i, chapter in enumerate(chapters):
          if chapter != None:
            if chapter_list[i] != None:
              chapter_list[i] = chapter_list[i] + self.__get_chapter_tuple(chapter)
            else:
              chapter_list[i] = self.__get_chapter_tuple(chapter)
      else: # key_chapter != None
        prev_chapter_list = key.get(current_key_chapter)
        current_key_chapter = self.__get_chapter_tuple(key_chapter) 
        for i, chapter in enumerate(chapters):
          if chapter == None:
            chapter_list.append(prev_chapter_list[i])
          elif chapter == "Empty": # NOTE If the value denoting an empty cell changes, change this to correspond to that
            chapter_list.append(None)
          else:
            chapter_list.append(self.__get_chapter_tuple(chapter))
      key.update({current_key_chapter:chapter_list})
    return (textbook_titles, key)

  """
  chapter: chapter title to create a tuple from; must NOT be None and should be formatted: chapter_num "chapter_title"
    where chapter_num is the chapter number/sheet name for that corresponding chapter and chapter_title is the actual chapter
    title from the textbook which MUST be enclosed in double quotes
  Takes the properly formatted chapter string and returns a tuple: (chapter_num, chapter_title)
  """
  def __get_chapter_tuple(self, chapter: str) -> tuple[str, str]:
    open_quote_index = chapter.find("\"")
    chapter_num = chapter[:open_quote_index].strip()
    chapter_title = chapter[open_quote_index+1:len(chapter)-1]
    return (chapter_num, chapter_title)
    