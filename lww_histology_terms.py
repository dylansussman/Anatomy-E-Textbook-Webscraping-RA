import re
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from striprtf.striprtf import rtf_to_text

DRIVER_PATH = '/usr/local/bin/chromedriver'
driver = webdriver.Chrome(executable_path=DRIVER_PATH)
# Link to go to first texbook to scrape: Atlas of Histology with Functional Correlations, 13e
driver.get('https://meded-lwwhealthlibrary-com.proxy.lib.ohio-state.edu/book.aspx?bookid=2992')

# Read in username and password for login
username: str = ''
password: str = ''
with open('../Documents/Private/login_credentials.rtf', 'r') as login:
    while username == '':
        username = rtf_to_text(login.readline())
    username = username[:username.find('/')]
    password = rtf_to_text(login.readline())

# Get thru login page
username_element: WebElement = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'username')))
password_element: WebElement = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'password')))
username_element.send_keys(username)
password_element.send_keys(password)
driver.find_element(By.ID, 'submit').click()

# Get to first chapter to scrape
chapter_1: WebElement = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, 'Chapter 1: Histologic Methods')))
chapter_1.click()

# Get chapter title
chapter_title: str = driver.find_element(By.CLASS_NAME, 'chapter-title').text

# Get headers for each chapter sub-section
headers: list[WebElement] = driver.find_elements(By.CLASS_NAME, 'scrollTo')
section_ids: list[str] = []
for element in headers:
    href: str = element.get_attribute('href')
    pound_index: int = href.find('#')
    id: str = f'section_{href[pound_index + 1:]}'
    section_ids.append(id)

# Creart 2D list with all bolded terms
keywords: list[list[str]] = []
for id in section_ids:
    # path: str = f"//div[@id='{id}']/descendant::div[not(@class='caption-legend')]/div[@class='para']/strong" # | //div[@id='{id}']/descendant::div[not(@class='caption-legend')]/ul[@class='bullet']/li/div[@class='para']/strong"
    path: str = f"//div[@id='{id}']/descendant::div[@class='para']/strong"
    bold_web_elements: list[WebElement] = driver.find_elements(By.XPATH, path)
    bold_words: list[str] = []
    for element in bold_web_elements:
        if len(element.text) > 1:
            bold_words.append(element.text)
    keywords.append(bold_words)

# Create list of headers (just text)
headers_text: list[str] = []
for header in headers:
    headers_text.append(header.accessible_name)

# Write everything to Excel file
wb: Workbook = Workbook()
title: str = re.sub(INVALID_TITLE_REGEX, '', chapter_title)
ws: Worksheet = wb.active
ws.title = title
col: int = 1
for header in headers_text:
    ws.cell(1, col, header)
    col += 1

row: int = 2
col = 1
for keyword_list in keywords:
    for keyword in keyword_list:
        ws.cell(row, col, keyword)
        row += 1
    row = 2
    col += 1

wb.save('../OneDrive - The Ohio State University/Survey Development - Dylan/Atlas of Histology with Functional Correlations, 13e.xlsx')
wb.close()



