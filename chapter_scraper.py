from selenium.webdriver.remote.webelement import WebElement
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Object to represent a chapter of a textbook
# Big picture: chapter_scraper is within book_scraper
class chapterScraper:
    BAD_HEADERS: list[str] = ['Summary', 'Review', 'Additional Histologic Images', 'Outline', 'KEY WORDS', 'Key words']
    ROMAN_NUMERALS: list[str] = ['I', 'V', 'X', 'L', 'C', 'D']
    
    def __init__(self, title: str, web_driver: webdriver.Chrome ) -> None:
        self.chapter_title = title
        self.driver = web_driver

    def get_headers(self) -> dict[str, WebElement]:
        headers: dict[str, WebElement] = {}
        # NOTE For scraping LWW Textbooks
        # sections: list[WebElement] = self.driver.find_elements(By.CLASS_NAME, 'scrollTo')
        # NOTE For scraping Elsevier Textbooks
        WebDriverWait(self.driver, 30).until(EC.visibility_of_any_elements_located((By.CSS_SELECTOR, 'a.c-link--nav')))
        # NOTE When using XPATH "//a" the headers are found
        sections: list[WebElement] = self.driver.find_elements(By.CSS_SELECTOR, "a.c-link--nav")
        for section in sections:
            if not any(map(lambda header: header in section.text, self.BAD_HEADERS)):
                # NOTE For scraping LWW Textbooks
                # href: str = section.get_attribute('href')
                # pound_index: int = href.find('#')
                # id: str = f'section_{href[pound_index + 1:]}'
                # NOTE For scraping Elsevier Textbooks
                id: str = section.get_attribute('scroll-to-id')
                if id != None:
                    headers.update({id:section})
        return headers

   
    def map_helper(self, word: WebElement):
        text = word.text.strip()
        if len(text) > 2 and not("(a)" in text or "(a, b)" in text or "EM " in text or "H&E" in text or "(HP)" in text or "(LP)" in text):
            if text[len(text) - 1] == '.':
                text = text[:text.find('.')].strip()
            return text
    
    # Only for scraping Elsevier textbooks
    # Appends figures' bold terms to header_term_dict
    def get_figure_terms(self, section_id: str, header_term_dict: dict[str, list[str]]) -> None:
        path: str = f"//section[@id='{section_id}']/div[@class='inline-image figure']"
        figures: list[WebElement] = self.driver.find_elements(By.XPATH, f"{path}")
        path += f"/figure/div[@class='caption-holder']"
        figure_titles: list[WebElement] = self.driver.find_elements(By.XPATH, f"{path}/div[1]")
        figure_ids: list[str] = list(map(lambda e: e.get_attribute("id"), figures))        
        for x, figure in enumerate(figures):
            bold_terms: list[str] = []
            path = f"//section[@id='{section_id}']/div[@id='{figure_ids[x]}']/figure/div[@class='caption-holder']"
            length: int = len(figure.find_elements(By.XPATH, f"{path}/div[@class='inline-image-caption']"))
            for i in range(2, length + 1):
                bold_terms += list(map(self.map_helper, figure.find_elements(By.XPATH, f"{path}/div[{i}]/i | {path}/div[{i}]/b")))
            if len(bold_terms) > 0 and any(list(map(lambda item: item != None, bold_terms))):
                paren_index: int = figure_titles[x].text.find(' (')
                title: str = figure_titles[x].text if paren_index == -1 else figure_titles[x].text[:paren_index]
                title = "Figure. " + title
                header_term_dict.update({title:bold_terms})

    def get_section_bold_terms(self, path: str) -> list[str]:
        bold_web_elements: list[WebElement] = self.driver.find_elements(By.XPATH, path)
        bold_words: list[str] = []
        for element in bold_web_elements:
            word = element.text.strip()
            if ':' in word:
                word = word[:word.find(':')]
            # NOTE For scraping LWW Textbooks
            # if len(word) > 1 and self.add_with_number(word) and not '.' in word and not self.is_roman_numeral(word) and word != "Image:":
            # NOTE For scraping Elsevier Textbooks
            if len(word) > 1 and self.add_with_number(word) and not self.is_roman_numeral(word) and word != "Image:":
                if not word.lower() in bold_words:
                    if word[len(word) - 1] == '.':
                        word = word[:word.find('.')].strip()
                    bold_words.append(word.lower())
        return bold_words

    def create_worksheet(self, ws: Worksheet, data: dict[str, list[str]]) -> None:
        # Add chapter name to first row of sheet and merge cells
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(data))
        ws.cell(row=1, column=1, value=self.chapter_title).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=1).fill = PatternFill('solid', fgColor='BFBFBF')
        row: int = 2
        col: int = 1
        max_width: int = 0
        for header, terms in data.items():
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            max_width = len(header)
            terms.sort(key=lambda t: len(t) if t != None else 0) # sort terms by length of term (string)
            for term in terms:
                if term:
                    row += 1
                    ws.cell(row=row, column=col, value=term)
                    if len(term) > max_width:
                        max_width = len(term)
            ws.column_dimensions[get_column_letter(col)].width = max_width
            row = 2
            col += 1

    def is_roman_numeral(self, word: str) -> bool:
        word_chars_bool: list[bool] = []
        for char in word:
            word_chars_bool.append(char in self.ROMAN_NUMERALS)
        return all(word_chars_bool)

    def add_with_number(self, word: str) -> bool:
        words_chars_bool: list[bool] = []
        for char in word:
            words_chars_bool.append(char.isnumeric())
        if any(words_chars_bool):
            if len(words_chars_bool) < 3:
                return False
            else:
                return True
        return True
            

