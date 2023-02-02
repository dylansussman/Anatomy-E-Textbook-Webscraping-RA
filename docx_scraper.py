from docx import Document
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import string

class docxScraper:
  def __init__(self) -> None:
    pass

  def open_file(self, file_name: str) -> Document:
    return Document(file_name)

  def get_file_bold_terms(self, doc: Document, chapter_dict: dict[str, list[str]]):
    title_found: bool = False
    chapter_title: str = ''
    bold_terms: list[str] = []
    for para in doc.paragraphs:
      if (not title_found and "Chapter" in para.text):
        chapter_title = para.text
        title_found = True
      elif (title_found and len(para.runs) > 0 and para.style.name == "Body Text"):
        # TODO Eliminate duplicates
        # TODO Ensure terms are > 1 character to add to bold list
        # TODO hyphen are replaced with \xad, so eliminate this at the end of words that have it
        for run in para.runs:
          if (run.text != '' and run.text != ' ' and run.bold):
            term: str = run.text
            term = term.strip().translate(str.maketrans('', '', string.punctuation)) # Delete leading and trailing whitespace and punctuation
            if len(term) > 1 and not term.lower() in bold_terms:
              # Eliminate hyphens at the end of words
              if ('\xad' in term):
                term = term[:term.find('\\')]
              bold_terms.append(term.lower())
    chapter_dict.update({chapter_title:bold_terms}) if len(bold_terms) > 0 else 0

  # Likely need equivalent of create_worksheet function in chapter_scraper and the workbook creation can just be done in main
  def create_worksheet(self, ws: Worksheet, data: list[str], chapter_title: str) -> None:
    # Add chapter name to first row of sheet and merge cells
    ws.cell(row=1, column=1, value=chapter_title).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=1).fill = PatternFill('solid', fgColor='BFBFBF')
    row: int = 1
    max_width: int = len(chapter_title)
    data.sort(key=lambda t: len(t) if t != None else 0) # sort terms by length of term (string)
    for term in data:
      if term:
          row += 1
          ws.cell(row=row, column=1, value=term)
          if len(term) > max_width:
              max_width = len(term)
    ws.column_dimensions[get_column_letter(1)].width = max_width