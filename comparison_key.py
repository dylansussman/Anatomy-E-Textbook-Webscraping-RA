from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

"""
This class represents the bank of words that will be tested against.
An instantiation of this class will contain a dictionary with every term from the textbook designated
as the key to compare terms from other textbooks against. These terms will be orgnanized by chapter
and this class contains methods to compare terms from other textbooks against the key
"""
class comparisonKey:
  """
  file_name: the name of the sheet designated as the key
  """
  def __init__(self, file_name: str) -> None:
    # Class variable to hold all terms from sheet designated as the comparison key
    wb: Workbook = load_workbook(file_name)
    self.key = self.__intialize_key(wb)

  """
  file_name: the name of the sheet designated as the key
  Reads the sheets from wb to create the key to compare against
  """
  def __intialize_key(self, wb: Workbook) -> dict[(str, str), list[str]]:
    key: dict[(str, str), list[str]] = {}
    for sheet_name in wb.sheetnames:
      sheet: Worksheet = wb[sheet_name]
      chapter_name: str = sheet['A1'].value
      terms: list[str] = []
      for value in sheet.iter_cols(min_row=3, values_only=True):
        for term in value:
          if term != None:
            terms.append(term)
      key.update({(sheet_name, chapter_name):terms})
    return key

    