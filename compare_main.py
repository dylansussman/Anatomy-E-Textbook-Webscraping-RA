from comparison_mappings import comparisonMappings
from comparison_key import comparisonKey
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

"""
NOTE
Should only ever need to change constants
If a new comparison is run with a different textbook as the dictionary of terms/key
  make sure to change OUTPUT_SHEET_NAME so that the textbook being used as the key is in parentheses
  to make sure it doesn't override an existing output sheet from another textbook as the key
"""
FOLDER_PATH = "../OneDrive - The Ohio State University/Survey Development - Dylan/Textbooks Data/"
OUTPUT_SHEET_NAME = "Term Comparison Output (Junqueira's).xlsx"
COMPARISON_SHEET_NAME = "Textbook Chapter Comparisons.xlsx"

mappings: comparisonMappings = comparisonMappings(f"{FOLDER_PATH}{COMPARISON_SHEET_NAME}")
key: comparisonKey = comparisonKey(f"{FOLDER_PATH}{mappings.key_textbook}.xlsx")

"""
NOTE
word_count_tracker: 
  key -> (general chapter name, term)
  value -> 8-item list of 0s, 1s where each index corresponds to the textbook of the same index in mappings.textbook_titles
  A 0 at an index indicates that that term did not appear in the corresponding textbook, and a 1 indicates that it did
"""
word_count_tracker: dict[tuple[str, str], list[int]] = {}
for k, v in mappings.chapters_key.items():
  for i, chapter in enumerate(v):
    if (chapter != None):
      while len(chapter) > 0: # Some chapter tuples have more than 1 chapter in them so need to iterate thru them all
        sheet_name, chapter_name, *chapter = chapter
        word_count_tracker = key.compare_textbook_chapter(FOLDER_PATH, k, mappings.textbook_titles, mappings.general_chapters.get(k), i, (sheet_name, chapter_name), word_count_tracker)

# Create workbook for output and write column headers for each sheet
wb: Workbook = Workbook()
wb.remove(wb.active)
temp_sheet_names = mappings.general_chapters.values()
sheet_names: list[str] = []
for sheet_name in temp_sheet_names:
  if not sheet_name in sheet_names:
    sheet_names.append(sheet_name)

for sheet in sheet_names:
  ws: Worksheet = wb.create_sheet(sheet)
  header_font: Font = Font(bold=True, size = 14)
  ws.cell(row=1, column=1, value="Term").font = header_font
  ws.cell(row=1, column=2, value="Count").font = header_font
  ws.cell(row=1, column=3, value=key.key_textbook_name).font = header_font
  ws.column_dimensions[get_column_letter(3)].width = len(key.key_textbook_name)
  for i in range(4, 4 + len(mappings.textbook_titles)):
    ws.cell(row=1, column=i, value=mappings.textbook_titles[i-4]).font = header_font
    ws.column_dimensions[get_column_letter(i)].width = len(mappings.textbook_titles[i-4])

wb.save(f"{FOLDER_PATH}{OUTPUT_SHEET_NAME}")
wb.close()

# Write out to sheets
wb = load_workbook(f"{FOLDER_PATH}{OUTPUT_SHEET_NAME}")
thin_side: Side = Side(border_style="thin", color="d4d4d4")
cell_border: Border = Border(top=thin_side, bottom=thin_side, right=thin_side, left=thin_side)
yellow_fill: PatternFill = PatternFill(fill_type="solid", start_color="00FFFF00")
for k, v in word_count_tracker.items():
  sheet_name, term = k
  ws = wb[sheet_name]
  r = ws.max_row + 1
  ws.cell(row=r, column=1, value=term)
  term_count = sum(v)
  ws.cell(row=r, column=2, value=term_count)
  ws.cell(row=r, column=3).fill = yellow_fill
  ws.cell(row=r, column=3).border = cell_border
  for i, count in enumerate(v):
    if count > 0:
      ws.cell(row=r, column=i + 4).fill = yellow_fill
      ws.cell(row=r, column=i + 4).border = cell_border

for sheet in wb.sheetnames:
  ws = wb[sheet]
  max_width = 0
  for term in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    term, = term
    if len(term) > max_width:
      max_width = len(term)
  ws.column_dimensions[get_column_letter(1)].width = max_width

wb.save(f"{FOLDER_PATH}{OUTPUT_SHEET_NAME}")
wb.close()